from openpyxl import Workbook
from selenium import webdriver
import time
from selenium.webdriver.common.keys import Keys
from bs4 import BeautifulSoup
from openpyxl import Workbook
from selenium.webdriver.common.by import By
from datetime import datetime
from django.http import HttpResponse

driver = webdriver.Chrome()

def scrape_vatia_com_co_tarifas():
    driver.get("https://vatia.com.co/tarifas-costo-unitario-mercado-regulado/")
    time.sleep(5)
    all_data = {}
    html_source = driver.page_source
    soup = BeautifulSoup(html_source, 'html.parser')

    try:
        celdas = soup.find('table').find('tbody').find_all('td')
        for i in range(0, len(celdas), 17):
            row_data = celdas[i:i+17]
            data_list = [cell.text.strip() for cell in row_data]
            if data_list[0].startswith('202402'):
                key = data_list[1] 
                filtered_data = data_list[1:3] + data_list[4:11]
                if key in all_data:
                    all_data[key].append(filtered_data)
                else:
                    all_data[key] = [filtered_data]

        for key, data_list in all_data.items():
            for data_row in data_list:
                data_row[0], data_row[1], data_row[2], data_row[3], data_row[4], data_row[5], data_row[6], data_row[7], data_row[8] = data_row[0], data_row[1], data_row[2], data_row[5], data_row[6], data_row[4], data_row[3], data_row[7], data_row[8] 
    
    except Exception as e:
        print("Error durante la extracción de datos:", e)
    print(all_data)
    return all_data

def scrape_bia_com_co_tarifas():
    driver.get("https://www.bia.app/tarifas")
    time.sleep(1)
    dropdown_buttons = driver.find_elements(By.CLASS_NAME, 'tarifas-table-filter-item-button-toggle')
    dropdown_buttons[1].click()

    time.sleep(0.5)
    all_data=[]
    company_elements = driver.find_elements(By.CLASS_NAME, 'tarifas-table-filter-item-dropdown-list-market-item')
    for index, company_element in enumerate(company_elements):
        empresa_nombre = company_element.find_element(By.TAG_NAME, 'div').text

        company_element.click()
        time.sleep(0.5)

        updated_html = driver.page_source
        dropdown_buttons = driver.find_elements(By.CLASS_NAME, 'tarifas-table-filter-item-button-toggle')
        dropdown_buttons[1].click()
        time.sleep(0.5)

        soup = BeautifulSoup(updated_html, 'html.parser')

        empresas = soup.find_all('div', class_='w-layout-grid tarifas-table-list-wrap')

        for empresa in empresas:
            
            celdas = empresa.find_all('div', class_='tarifas-table-list-column-item')[0:5]+empresa.find_all('div', class_='tarifas-table-list-column-item')[6:16] + empresa.find_all('div', class_='tarifas-table-list-column-item')[21:46]

            datos_num=[]
            for celda in celdas:
                data_list=celda.find("div").text.strip()
                datos_num.append(data_list)
            
            reorganizado = [empresa_nombre, datos_num[0], datos_num[5], datos_num[15], datos_num[20], datos_num[10], datos_num[25], datos_num[30], datos_num[35],
            empresa_nombre,datos_num[1], datos_num[6], datos_num[16], datos_num[21], datos_num[11], datos_num[26], datos_num[31],datos_num[36],
            empresa_nombre,datos_num[2], datos_num[7], datos_num[17], datos_num[22], datos_num[12], datos_num[27], datos_num[32],datos_num[37],
            empresa_nombre,datos_num[3], datos_num[8], datos_num[18], datos_num[23], datos_num[13], datos_num[28], datos_num[33],datos_num[38],
            empresa_nombre,datos_num[4], datos_num[9], datos_num[19], datos_num[24], datos_num[14], datos_num[29], datos_num[34],datos_num[39]]
            subarrays = [reorganizado[i:i+9] for i in range(0, len(reorganizado), 9)]
            all_data.extend(subarrays)

        time.sleep(0.5)
    return all_data

def scrape_neu_com_co_tarifas():

    driver.get("https://www.neu.com.co/tarifas")
    time.sleep(1)

    dropdown_button = driver.find_element(By.CLASS_NAME, 'rs-picker-toggle')
    dropdown_button.click()
    time.sleep(0.5)
    all_data=[]
  
    company_elements = driver.find_elements(By.CLASS_NAME, 'rs-picker-select-menu-item')
    for index, company_element in enumerate(company_elements):
        for _ in range(1):
            dropdown_button.send_keys(Keys.ARROW_DOWN)
            time.sleep(1)

        dropdown_button.send_keys(Keys.ENTER)
        time.sleep(2)

        updated_html = driver.page_source

        dropdown_button = driver.find_element(By.CLASS_NAME, 'rs-picker-toggle')
        dropdown_button.click()
        empresa_nombre = driver.find_element(By.TAG_NAME, 'span').text
        time.sleep(2)

        # Utilizar BeautifulSoup para analizar el HTML
        soup = BeautifulSoup(updated_html, 'html.parser')

        # Encontrar la información de la tabla por empresa
        empresas = soup.find_all('div', class_='rs-table-row')
        for empresa in empresas:
            # Obtener el nombre de la empresa


            # Obtener la información de las celdas en la misma fila
            celdas = empresa.find_all('div', class_='rs-table-cell')
            # Imprimir la información de cada celda
            datos_num = []
            
            for celda in celdas:
                data_list = celda.find("div").text.strip()
                datos_num.append(data_list)
            
            if all(char.isalpha() for char in ''.join(datos_num)):
                continue

        # Reorganizar y agregar a all_data solo si datos_num no contiene solo letras
            reorganizado = [empresa_nombre, datos_num[0],datos_num[2], datos_num[3], datos_num[4], datos_num[6], datos_num[5], datos_num[7], datos_num[1]]
            all_data.append(reorganizado)
        time.sleep(0.5)
        
       
    driver.quit()
    return all_data


def data_to_excel(scraped_data_bia, scraped_data_neu, scraped_data_vatia):
    workbook = Workbook()
    sheet = workbook.active

    current_month_year = datetime.now().strftime("%B %Y")
    headers = [
        "PERIODO", "AGENTE", "MERCADO_COMERC", "NIVEL_TENSION",
        "TARIFA_G", "TARIFA_T", "TARIFA_D",
        "TARIFA_PR", "TARIFA_Cv", "TARIFA_R", "TARIFA_CUv", "TARIFA_CU_APL"
    ]

    sheet.append(headers)

    for i in range(2, sum(len(sublist) for sublist in scraped_data_vatia.values()) + len(scraped_data_bia)+len(scraped_data_neu) + 2):
        sheet.cell(row=i, column=1, value=current_month_year)
        if i < sum(len(sublist) for sublist in scraped_data_vatia.values()) + 2:
            sheet.cell(row=i, column=2, value="GNCC")
        elif i < sum(len(sublist) for sublist in scraped_data_vatia.values()) + len(scraped_data_bia)+2:
            sheet.cell(row=i, column=2, value="BIAC")
        else:
            sheet.cell(row=i, column=2, value="NEUC")

    row_index = 2
    for company, data_list in scraped_data_vatia.items():
        for data_row in data_list:
            for col_index, cell_value in enumerate(data_row, start=3):
                sheet.cell(row=row_index, column=col_index, value=cell_value)
            row_index += 1

    for row_index, row_values in enumerate(scraped_data_bia, start=sum(len(sublist) for sublist in scraped_data_vatia.values()) +2):
        for col_index, cell_value in enumerate(row_values, start=3):
            sheet.cell(row=row_index, column=col_index, value=cell_value)

    
    for row_index, row_values in enumerate(scraped_data_neu, start=sum(len(sublist) for sublist in scraped_data_vatia.values())+len(scraped_data_bia) + 2):
        for col_index, cell_value in enumerate(row_values, start=3):
            sheet.cell(row=row_index, column=col_index, value=cell_value)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    workbook.save(response)
    response['Content-Disposition'] = 'attachment; filename="Excel.xlsx"'
    return response


scraped_data_vatia = scrape_vatia_com_co_tarifas()
scraped_data_bia = scrape_bia_com_co_tarifas()
scraped_data_neu = scrape_neu_com_co_tarifas()
data_to_excel(scraped_data_bia, scraped_data_neu, scraped_data_vatia)


