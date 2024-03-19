import requests
import locale
from bs4 import BeautifulSoup
from openpyxl.styles import Font
from datetime import datetime
from openpyxl import Workbook
from io import BytesIO
import webbrowser
from pdfminer.high_level import extract_text
import re

url2 = 'https://www.celsia.com/es/informacion-regulatoria-y-res-creg-080/tarifas/'

def status_code_url(url):
    try:
        r = requests.get(url, timeout=10)
        r.raise_for_status() 
        s = BeautifulSoup(r.text, 'lxml')
        return s
    except requests.exceptions.RequestException as e:
        print(f'Error al obtener la página: {e}')
        return None

def get_current_month_link(s):
    locale.setlocale(locale.LC_TIME, 'es_ES.UTF-8')
    current_month = 'febrero'  # Puedes cambiar esto a datetime.now().strftime('%B') si prefieres el mes actual
    links = s.find('tbody').find_all('a')

    for link in links:
        if current_month.lower() in link.text.lower():
            href_value = link['href']
            
            # Verificar si el valor de href parece ser una URL válida
            if re.match(r'https?://\S+', href_value):
                return href_value
            else:
                print(f'El valor de href no es una URL válida para el mes {current_month}')
                return None

    print(f'No se encontró el enlace para el mes {current_month}')
    return None

def clean_text(text):
    return re.sub(r'[^\x20-\x7E]', '', text)

def convert_pdf_to_excel(pdf_url):
    pdf_response = requests.get(pdf_url)

    if pdf_response.status_code != 200:
        print(f'Error al descargar el PDF. Código de estado: {pdf_response.status_code}')
        return None

    pdf_text = extract_text(BytesIO(pdf_response.content))

    capturing_data = False
    keywords = ["(G)", "(D)", "(PR)", "(Cv)", "(CUv) iii"]
    transmision_keyword = "(T)"
    restricciones_keyword = "(R)"
    start_row = 2

    for keyword in keywords:
        col_num = headers.index(f"TARIFA_{keyword.strip(' ()iii,')}") + 1
        capturing_data = False
        data_list = []

        for line in pdf_text.split('\n'):
            cleaned_line = clean_text(line)

            if keyword in cleaned_line:
                capturing_data = True
            elif capturing_data and line.strip():
                data_list.append(cleaned_line)

                # Detener la captura después de 11 líneas
                if len(data_list) == 11:
                    capturing_data = False

        # Escribir los datos en la hoja de cálculo desde la fila 2
        for i in range(4):
            sheet.cell(row=start_row + i, column=col_num, value=data_list[i])

        for i in range(5, 9):
            sheet.cell(row=start_row + i - 1, column=col_num, value=data_list[i])

        for i in range(9, 11):
            sheet.cell(row=start_row + i - 2, column=col_num, value=data_list[i])

    # Para las keywords "(T)" y "(R)"
    for keyword in [transmision_keyword, restricciones_keyword]:
        col_num = headers.index(f"TARIFA_{keyword.strip(' ()iii,')}") + 1
        capturing_data = False
        data_list = []

        for line in pdf_text.split('\n'):
            cleaned_line = clean_text(line)

            if keyword in cleaned_line:
                capturing_data = True
            elif capturing_data and line.strip():
                data_list.append(cleaned_line)

                # Detener la captura después de 11 líneas
                if len(data_list) == 11:
                    capturing_data = False

        # Escribir el primer dato en las columnas "TARIFA_T" y "TARIFA_R"
        value_to_repeat = data_list[0]
        for i in range(9):
            sheet.cell(row=start_row + i, column=col_num, value=value_to_repeat)

    # Repetir la URL en la columna "AGENTE"
    col_num_agente = headers.index("AGENTE") + 1
    for i in range(9):
        sheet.cell(row=start_row + i, column=col_num_agente, value=url2)

    for column in sheet.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column[0].column_letter].width = adjusted_width

    print(f'Se ha creado el archivo Excel: {file_name}')
    return file_name


if __name__ == '__main__':
    s = status_code_url(url2)
    if s:
        current_month_link = get_current_month_link(s)
        if current_month_link:
            print(f'Enlace para el mes actual ({datetime.now().strftime("%B")}): {current_month_link}')

            pdf_table = convert_pdf_to_excel(current_month_link)