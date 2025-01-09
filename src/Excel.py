from openpyxl import Workbook
from selenium import webdriver
import time
from selenium.webdriver.common.keys import Keys
from bs4 import BeautifulSoup
from openpyxl import Workbook
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from datetime import datetime
from selenium.webdriver.support.ui import Select
import pandas as pd
import tabula
import requests
import re
import pdfplumber
import os


driver = webdriver.Chrome()



def scrape_ettc_com_co_tarifas():
    mes_actual = datetime.now().month
    print(mes_actual)
    mes_anterior = (datetime.now().replace(day=1) - pd.DateOffset(months=1)).month
    print(mes_anterior)
    # Nombres de los meses en español
    meses = ['Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio', 'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']
    mes_anterior_nombre = meses[mes_anterior - 1]
    url = "https://www.enertotalesp.com/soporte-en-linea/tarifas-publicadas/"
    response= requests.get(url)
    soup = BeautifulSoup(response.text, 'html.parser')
    link_abril = soup.find('a', text=mes_anterior_nombre)
    if link_abril:
        pdf_url = link_abril['href']
        print("Extrayendo datos del PDF...")

        pdf_response = requests.get(pdf_url)
        if pdf_response.status_code == 200:
            pdf_path = 'temp.pdf'
            with open(pdf_path, 'wb') as f:
                f.write(pdf_response.content)
            with pdfplumber.open(pdf_path) as pdf:
                tables = []
                for page in pdf.pages:
                    tables.extend(page.extract_tables())
            
            os.remove(pdf_path) 
            
            dataframes = [pd.DataFrame(table[1:], columns=table[0]) for table in tables]
            
            combined_df = pd.concat(dataframes, ignore_index=True)
            
            Tm = combined_df.iloc[2, 12]
            Rm = combined_df.iloc[2, 14]
            Tm = Tm.replace(',', '.')
            Rm = Rm.replace(',', '.')
            Tm = pd.to_numeric(Tm, errors='coerce')
            Rm = pd.to_numeric(Rm, errors='coerce')
            combined_df = combined_df.iloc[3:]
            combined_df = combined_df.iloc[:, 11:-6]
            combined_df = combined_df.applymap(lambda x: x.replace(',', '.') if isinstance(x, str) else x)
            combined_df = combined_df.apply(pd.to_numeric, errors='coerce')
            combined_df = combined_df.dropna()


            print(combined_df)
            combined_df['Tm'] = Tm
            combined_df['Rm'] = Rm
            combined_df.columns = ['G', 'C', 'Pr', 'D', 'Cu', 'Tm', 'Rm']
            combined_df = combined_df.reindex(columns=['G', 'Tm', 'D', 'Pr', 'C', 'Rm', 'Cu'])

            array_of_arrays = combined_df.iloc[1:].values.tolist()

            # Devolver la lista de listas
            return array_of_arrays
        else:
            print("Error al descargar el PDF:", pdf_response.status_code)
            return None
    else:
        print("No se ha publicado tarifas para dicho mes")

def scrape_rtqc_com_co_tarifas():
    mes_actual = datetime.now().month
    mes_anterior = (datetime.now().replace(day=1) - pd.DateOffset(months=1)).month
    
    # Nombres de los meses en español
    meses = ["ENERO", "FEBRERO", "MARZO", "ABRIL", "MAYO", "JUNIO", "JULIO", "AGOSTO", "SEPTIEMBRE", "OCTUBRE", "NOVIEMBRE", "DICIEMBRE"]
    mes_anterior_nombre = meses[mes_anterior - 1]
    url = "https://www.ruitoqueesp.com/nuevo/servicios/energia/"
    response = requests.get(url)
    soup = BeautifulSoup(response.text, 'html.parser')
    link_mayo = soup.find('a', text=mes_anterior_nombre)  
    if link_mayo:
        pdf_url = link_mayo['href']
        print("Extrayendo datos del PDF...")
        pdf_response = requests.get("https://www.ruitoqueesp.com/" + pdf_url)
        if pdf_response.status_code == 200:
            pdf_path = 'temp.pdf'
            with open(pdf_path, 'wb') as f:
                f.write(pdf_response.content)
            with pdfplumber.open(pdf_path) as pdf:
                first_page = pdf.pages[0]
                table = first_page.extract_table()
            
            os.remove(pdf_path)
            
            if not table:
                print("No se encontraron tablas relevantes en el PDF.")
                return None
            
            # Crear DataFrame
            combined_df = pd.DataFrame(table).iloc[3:, 1:9]
            print(combined_df)

            def extract_number(text):
                match = re.match(r'(\d+)', text)
                return int(match.group(1)) if match else None
            
            def clean_currency(x):
                if isinstance(x, str):
                    return x.replace('$', '').replace(' ', '')
                return x
            
            # Limpiar los datos de moneda
            for col in combined_df.columns[1:]:
                combined_df[col] = combined_df[col].map(clean_currency)
    
            combined_df.iloc[:, 1:] = combined_df.iloc[:, 1:].apply(pd.to_numeric, errors='coerce')
            
            # Renombrar columnas
            combined_df.columns = ['NT', 'G', 'T', 'D', 'Rm', 'C', 'Pr', 'Cu']

            combined_df['NT'] = combined_df['NT'].apply(extract_number)
            
            # Reordenar columnas
            combined_df = combined_df.reindex(columns=['NT','G', 'T', 'D', 'Pr', 'C', 'Rm', 'Cu']) 
            combined_df['Cu_Repeated'] = combined_df['Cu']
            
            data_array = combined_df.to_numpy()[1:]
            
            return data_array
        else:
            print("Error al descargar el PDF:", pdf_response.status_code)
            return None
    else:
        print("No se ha encontrado el enlace para el mes anterior")
        return None



def scrape_qia_com_co_tarifas():
    mes_actual = datetime.now().month
    print(mes_actual)
    mes_anterior = (datetime.now().replace(day=1) - pd.DateOffset(months=1)).month
    print(mes_anterior)
    # Nombres de los meses en español
    meses = ["ENERO", "FEBRERO", "MARZO", "ABRIL", "MAYO", "JUNIO", "JULIO", "AGOSTO", "SEPTIEMBRE", "OCTUBRE", "NOVIEMBRE", "DICIEMBRE"]
    mes_anterior_nombre = meses[mes_anterior - 1]
    print(mes_anterior_nombre)
    url = "https://qienergy.co/tarifas/" 
    response = requests.get(url)
    soup = BeautifulSoup(response.text, 'html.parser')
    link_mes_anterior = soup.find('a', text=mes_anterior_nombre)
    if link_mes_anterior:
        pdf_url = link_mes_anterior['href']
        print("Extrayendo datos del PDF...")
        response = requests.get(pdf_url)
        if response.status_code == 200:
            tables = tabula.read_pdf(pdf_url, pages='all', multiple_tables=True)
            rows = []
            for table in tables:
                for index, row in table.iterrows():
                    rows.append(row.tolist()[2:10])
            datos_pdf = rows 
        else:
            print("Error al descargar el PDF:", response.status_code)
            return None
        
        if datos_pdf is not None:
            print("Datos extraídos con éxito:")
            elemento_5 = datos_pdf[1][5]
            print(elemento_5)
            elemento_6 = datos_pdf[1][7]
            print(elemento_6)
            elemento_anterior = None 
            all_rows = [] 
            for index, row in enumerate(datos_pdf[3:99]):
                print(datos_pdf[1:99])
                row.insert(6, elemento_5)
                row.insert(7, elemento_6)
                if not pd.isna(row[0]): 
                    elemento_anterior = row[0]  
                else:
                    row[0] = elemento_anterior
                
                tercer_elemento = row[0]
                if tercer_elemento is not None:
                    numero = float(''.join(filter(str.isdigit, tercer_elemento)))
                    modified_row = [numero] 
                else:
                    numero = None  
                    modified_row = [numero]  
                for item in row[1:8]:
                    if isinstance(item, str): 
                        modified_row.append(float(item.replace(',', '.')))
                    else:
                        modified_row.append(item)
                all_rows.append(modified_row)
            
            organized_rows = []
            for row in all_rows:
                organized_row = [row[0], row[1], row[6], row[2],  row[4], row[3],  row[7],  row[5], row[5]]
                organized_rows.append(organized_row)
                print(organized_rows) 
                print(mes_actual)
                print(mes_anterior)
            return organized_rows 
        else:
            print("No se pudieron extraer datos del PDF.")
    else:
        print("No se encontró el enlace del mes")


def scrape_vatia_com_co_tarifas():
    # Obtener el mes y año actual
    ahora = datetime.now()
    mes_actual = ahora.month
    anio_actual = ahora.year
    mes = 12
    if ahora.month < 12:
        anio_actual -= 1
    ciclo_value = f"{anio_actual}{mes:02d}"
    print(ciclo_value)
    mapeo = {
        "BAJO PUTUMAYO": "EBPD",
        "ENERTOLIMA": "CTSD",
        "EBSA": "EBSD",
        "ELECTROHUILA": "HLAD",
        "ELECTROMETA": "EMSD",
        "ESSA": "ESSD",
        "ELECTROCAQUETA": "CQTD",
        "CENS": "CNSD",
        "CETSA": "CETD",
        "CEDENAR": "CDND",
        "CARIBEMAR": "CMMD",
        "CODENSA-EEC": "ENDD",
        "CARIBESOL": "CSSD",
        "EPM-EADE": "EPMD",
        "CEDELCA": "CDID",
        "CHEC": "CHCD",
        "ENERCA": "CASD",
        "EMCALI": "EMID",
        "EMCARTAGO": "CTID",
        "EDEQ": "EDQD",
        "ENELAR": "ENID",
        "EPSA": "EPSD",
        "PUTUMAYO": "EPTD",
        "EEP": "EEPD"
    }
    driver.get("https://vatia.com.co/tarifas-costo-unitario-mercado-regulado/")
    time.sleep(15)
   
    select_element = driver.find_element(By.ID, 'ciclo')
    select_dropdown = Select(select_element)
    select_dropdown.select_by_value(ciclo_value)
    
    
    time.sleep(2) 
    
    all_data = []
    html = driver.page_source
    soup = BeautifulSoup(html, 'html.parser')

    try:
        celdas = soup.find('table').find('tbody').find_all('td')
        for i in range(0, len(celdas), 18):
            row_data = celdas[i:i+18]
            data_list = [cell.text.strip() for cell in row_data]
            all_data.append(data_list)
        
        # Procesa los datos
        processed_data = []
        for data_list in all_data:
            if len(data_list) >= 18: 
                data_list[5] = mapeo.get(data_list[5], data_list[5])
                try:
                    processed_row = [  
                        data_list[5],  # Campo mapeado
                        data_list[6],
                        int(data_list[3]),
                        float(data_list[7]),
                        float(data_list[10]),
                        float(data_list[11]),
                        float(data_list[9]),
                        float(data_list[8]),
                        float(data_list[12]),
                        float(data_list[14]), 
                        float(data_list[14])
                    ]
                    processed_data.append(processed_row)
                except ValueError as e:
                    print(f"Error al procesar la fila: {data_list} - {e}")

    except Exception as e:
        print("Error durante la extracción de datos:", e)

    print(processed_data)
    return processed_data


def scrape_bia_com_co_tarifas():
    ahora = datetime.now()
    mes_actual = ahora.month
    anio_actual = ahora.year
    mes = 12
    if ahora.month < 12:
        anio_actual -= 1
    ciclo_value = f"01-{mes:02d}-{anio_actual}"

    driver.get("https://www.bia.app/tarifas")
    time.sleep(1)

    dropdown_buttons = driver.find_elements(By.CLASS_NAME, 'tarifas-table-filter-item-button-toggle')
    dropdown_buttons[0].click()
    time.sleep(1)
    
    slide_element = driver.find_element(By.ID, 'slider-arrow-right')
    slide_element.click()
    time.sleep(2)

    dates_element = driver.find_element(By.ID, ciclo_value)
    dates_element.click()
    time.sleep(1)
    
    dropdown_buttons[1].click()
    time.sleep(1)
    
    all_data = []
    company_elements = driver.find_elements(By.CLASS_NAME, 'tarifas-table-filter-item-dropdown-list-market-item')
    
    for index, company_element in enumerate(company_elements):
        time.sleep(2)
        company_element.click()
        time.sleep(1)

        updated_html = driver.page_source
        dropdown_buttons = driver.find_elements(By.CLASS_NAME, 'tarifas-table-filter-item-button-toggle')
        dropdown_buttons[1].click()
        time.sleep(0.5)

        soup = BeautifulSoup(updated_html, 'html.parser')

        empresas = soup.find_all('div', class_='w-layout-grid tarifas-table-list-wrap')

        for empresa in empresas:
            
            celdas = empresa.find_all('div', class_='tarifas-table-list-column-item')[0:5]+empresa.find_all('div', class_='tarifas-table-list-column-item')[6:16] + empresa.find_all('div', class_='tarifas-table-list-column-item')[21:46]

            datos_num=[]
            for celda in celdas:
                data_list=celda.find("div").text.strip()
                datos_num.append(data_list)
            
            reorganizado = [int(re.search(r'\d+',datos_num[0]).group()), float(datos_num[5].replace(',','.')), float(datos_num[15].replace(',','.')), float(datos_num[20].replace(',','.')), float(datos_num[25].replace(',','.')), float(datos_num[10].replace(',','.')), float(datos_num[30].replace(',','.')), float(datos_num[35].replace(',','.')),float(datos_num[35].replace(',','.')),
            int(re.search(r'\d+',datos_num[1]).group()), float(datos_num[6].replace(',','.')), float(datos_num[16].replace(',','.')), float(datos_num[21].replace(',','.')), float(datos_num[26].replace(',','.')), float(datos_num[11].replace(',','.')), float(datos_num[31].replace(',','.')),float(datos_num[36].replace(',','.')),float(datos_num[36].replace(',','.')),
            int(re.search(r'\d+',datos_num[2]).group()), float(datos_num[7].replace(',','.')), float(datos_num[17].replace(',','.')), float(datos_num[22].replace(',','.')), float(datos_num[27].replace(',','.')), float(datos_num[12].replace(',','.')), float(datos_num[32].replace(',','.')),float(datos_num[37].replace(',','.')),float(datos_num[37].replace(',','.')),
            int(re.search(r'\d+',datos_num[3]).group()), float(datos_num[8].replace(',','.')), float(datos_num[18].replace(',','.')), float(datos_num[23].replace(',','.')), float(datos_num[28].replace(',','.')), float(datos_num[13].replace(',','.')), float(datos_num[33].replace(',','.')),float(datos_num[38].replace(',','.')),float(datos_num[38].replace(',','.')),
            int(re.search(r'\d+',datos_num[4]).group()), float(datos_num[9].replace(',','.')), float(datos_num[19].replace(',','.')), float(datos_num[24].replace(',','.')), float(datos_num[29].replace(',','.')), float(datos_num[14].replace(',','.')), float(datos_num[34].replace(',','.')),float(datos_num[39].replace(',','.')),float(datos_num[39].replace(',','.'))]
            subarrays = [reorganizado[i:i+9] for i in range(0, len(reorganizado), 9)]
            all_data.extend(subarrays)

        time.sleep(1)
        print(all_data)
    return all_data
    
def scrape_neu_com_co_tarifas():
    driver.get("https://erco.energy/co/servicios/comercializacion/tarifas")
    time.sleep(1)
    date_input = driver.find_element(By.CLASS_NAME, 'rs-input')
    time.sleep(10)
    dropdown_button = driver.find_element(By.CLASS_NAME, 'rs-picker-toggle')
    time.sleep(5)
    dropdown_button.click()
    time.sleep(5)
    all_data = []
    company_elements = driver.find_elements(By.CLASS_NAME, 'rs-picker-select-menu-item')
    
    additional_values = [1, 1, 1, 2, 3]
    value_index = 0

    for index, company_element in enumerate(company_elements):
        for _ in range(1):
            dropdown_button.send_keys(Keys.ARROW_DOWN)
            time.sleep(2)

        dropdown_button.send_keys(Keys.ENTER)
        time.sleep(4)

        updated_html = driver.page_source
        time.sleep(4)
        dropdown_button = driver.find_element(By.CLASS_NAME, 'rs-picker-toggle')
        time.sleep(4)
        dropdown_button.click()
        empresa_nombre = driver.find_element(By.TAG_NAME, 'span').text
        time.sleep(4)
        soup = BeautifulSoup(updated_html, 'html.parser')

        empresas = soup.find_all('div', class_='rs-table-row')
        datos_empresa = []
        for empresa in empresas:
            celdas = empresa.find_all('div', class_='rs-table-cell')
            datos_num = []

            for celda in celdas:
                data_list = celda.find("div").text.strip()
                datos_num.append(data_list)
            if all(char.isalpha() for char in ''.join(datos_num)):
                continue
            print(datos_num)
            if len(datos_num) >= 9:  
                filtered_data = [ 
                    datos_num[1],  # G
                    datos_num[2],  # T
                    datos_num[3],  # D
                    datos_num[4],  # C
                    datos_num[5],  # PR
                    datos_num[6],   # R
                    datos_num[7]  # CU
                ]
                datos_empresa.append(filtered_data)

            elif len(datos_num) >= 8:
                filtered_data = [ 
                     # CU
                    datos_num[1],  # G
                    datos_num[2],  # T
                    datos_num[3],  # D
                    datos_num[4],  # C
                    datos_num[5],  # PR
                    datos_num[6],
                    datos_num[7]    # R
                ]
                datos_empresa.append(filtered_data)
    
        for datos in datos_empresa:
            try:
                # Comprobación de longitud de la lista
                if len(datos) < 8:
                    print(f"Datos incompletos encontrados: {datos}")
                    continue
                
                # Convertir los valores a flotantes de forma segura
                reorganizado = [
                    additional_values[value_index],     # Valor adicional
                    float(datos[1].replace(',', '.') if ',' in datos[1] else datos[1]),  # G
                    float(datos[2].replace(',', '.') if ',' in datos[2] else datos[2]),  # T
                    float(datos[3].replace(',', '.') if ',' in datos[3] else datos[3]),  # D
                    float(datos[4].replace(',', '.') if ',' in datos[4] else datos[4]),  # PR
                    float(datos[5].replace(',', '.') if ',' in datos[5] else datos[5]),  # C
                    float(datos[6].replace(',', '.') if ',' in datos[6] else datos[6]),  # R
                    float(datos[7].replace(',', '.') if ',' in datos[7] else datos[7]),  # CU
                    float(datos[7].replace(',', '.') if ',' in datos[7] else datos[7])   # CU (duplicado)
                ]
                all_data.append(reorganizado)
                value_index = (value_index + 1) % len(additional_values)  # Actualizar el índice
            except ValueError as e:
                print(f"Error converting data to float: {e}")
        time.sleep(1)

    print(all_data)
    return all_data
def data_to_excel(scraped_data_ettc, scraped_data_neu, scraped_data_bia, scraped_data_vatia, scraped_data_qia):

    workbook = Workbook()
    sheet = workbook.active
    current_month_year = datetime.now().strftime("%B %Y")
    headers = [
        "PERIODO", "ID_AGENTE", "ID_MERCADO", "PROPIEDAD_ACTIVOS", "NIVEL_TENSION",
        "TARIFA_G", "TARIFA_T", "TARIFA_D",
        "TARIFA_PR", "TARIFA_Cv", "TARIFA_R", "TARIFA_CUv", "TARIFA_CU_APL", "COMPROBACION_CU", "COMPROBACION_T","COMPROBACION_R", 
        "COMPROBACION_D"
    ]
    sheet.append(headers)

    current_date = datetime.now()
    previous_month = current_date.month - 1 if current_date.month > 1 else 12
    previous_year = current_date.year if current_date.month > 1 else current_date.year - 1
    formatted_date = f"{previous_year}{previous_month:02d}"

    meses = ["ENERO", "FEBRERO", "MARZO", "ABRIL", "MAYO", "JUNIO", "JULIO", "AGOSTO", "SEPTIEMBRE", "OCTUBRE", "NOVIEMBRE", "DICIEMBRE"]
    previous_month_name = meses[previous_month - 1]

    for i in range(2, len(scraped_data_ettc)+ len(scraped_data_qia) + len(scraped_data_bia) + len(scraped_data_vatia) + len(scraped_data_neu)+2):
        sheet.cell(row=i, column=1, value=formatted_date)
        if i < len(scraped_data_qia) + 2:
            sheet.cell(row=i, column=2, value="QIEC")
        elif i < len(scraped_data_qia) +len(scraped_data_neu) + 2:
            sheet.cell(row=i, column=2, value="NEUC")
        elif i < len(scraped_data_qia) + len(scraped_data_neu)+ len(scraped_data_bia) +2:
            sheet.cell(row=i, column=2, value="BIAC")
        elif i < len(scraped_data_qia) + len(scraped_data_neu)+ len(scraped_data_bia) +len(scraped_data_vatia)+2: 
            sheet.cell(row=i, column=2, value="GNCC")
        elif i < len(scraped_data_qia) + len(scraped_data_neu)+ len(scraped_data_bia) +len(scraped_data_vatia)+2+len(scraped_data_ettc): 
            sheet.cell(row=i, column=2, value="ETTC")

    for row_index, row_values in enumerate(scraped_data_qia, start=2):
        for col_index, cell_value in enumerate(row_values, start=5):
            sheet.cell(row=row_index, column=col_index, value=cell_value)

    for row_index, row_values in enumerate(scraped_data_neu, start=len(scraped_data_qia)+2):
        for col_index, cell_value in enumerate(row_values, start=5):
            sheet.cell(row=row_index, column=col_index, value=cell_value) 

    for row_index, row_values in enumerate(scraped_data_bia, start=len(scraped_data_neu)+len(scraped_data_qia) +2):
        for col_index, cell_value in enumerate(row_values, start=5):
            sheet.cell(row=row_index, column=col_index, value=cell_value)

    for row_index, row_values in enumerate(scraped_data_vatia, start=len(scraped_data_neu)+len(scraped_data_qia)+len(scraped_data_bia) +2):
        for col_index, cell_value in enumerate(row_values, start=3):
            sheet.cell(row=row_index, column=col_index, value=cell_value)
            
    for row_index, row_values in enumerate(scraped_data_ettc, start=len(scraped_data_vatia)+len(scraped_data_neu)+len(scraped_data_qia)+len(scraped_data_bia) +2):
        for col_index, cell_value in enumerate(row_values, start=5):
            sheet.cell(row=row_index, column=col_index, value=cell_value)

    workbook.save(f"{previous_month_name}.xlsx")
    driver.quit()

scraped_data_qia = scrape_qia_com_co_tarifas()
scraped_data_neu=scrape_neu_com_co_tarifas() 
scraped_data_bia = scrape_bia_com_co_tarifas()
scraped_data_vatia = scrape_vatia_com_co_tarifas()
scraped_data_ettc=scrape_ettc_com_co_tarifas()

data_to_excel(scraped_data_ettc,  scraped_data_neu, scraped_data_bia, scraped_data_vatia, scraped_data_qia )


