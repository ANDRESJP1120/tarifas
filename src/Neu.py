import time
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from bs4 import BeautifulSoup
from openpyxl import Workbook
from selenium.webdriver.common.by import By
from datetime import datetime

driver = webdriver.Chrome()

def scrape_neu_com_co_tarifas():
    # Abrir el sitio web
    driver.get("https://www.neu.com.co/tarifas")
    time.sleep(1)

    # Hacer clic en el botón para desplegar todas las empresas
    dropdown_button = driver.find_element(By.CLASS_NAME, 'rs-picker-toggle')
    dropdown_button.click()
    time.sleep(0.5)
    all_data=[]
    # Encontrar y hacer clic en cada empresa utilizando teclas de flecha y Enter
    company_elements = driver.find_elements(By.CLASS_NAME, 'rs-picker-select-menu-item')
    for index, company_element in enumerate(company_elements):
        for _ in range(1):
            dropdown_button.send_keys(Keys.ARROW_DOWN)
            time.sleep(0.5)

        # Presionar "Enter" para seleccionar la empresa
        dropdown_button.send_keys(Keys.ENTER)
        time.sleep(0.5)

        updated_html = driver.page_source

        dropdown_button = driver.find_element(By.CLASS_NAME, 'rs-picker-toggle')
        dropdown_button.click()
        empresa_nombre = driver.find_element(By.TAG_NAME, 'span').text
        time.sleep(0.5)

        # Utilizar BeautifulSoup para analizar el HTML
        soup = BeautifulSoup(updated_html, 'html.parser')

        # Encontrar la información de la tabla por empresa
        empresas = soup.find_all('div', class_='rs-table-row')
        for empresa in empresas:
            # Obtener el nombre de la empresa


            # Obtener la información de las celdas en la misma fila
            celdas = empresa.find_all('div', class_='rs-table-cell')
            # Imprimir la información de cada celda
            datos_num = []
            
            for celda in celdas:
                data_list = celda.find("div").text.strip()
                datos_num.append(data_list)
            
            if all(char.isalpha() for char in ''.join(datos_num)):
                continue

        # Reorganizar y agregar a all_data solo si datos_num no contiene solo letras
            reorganizado = [empresa_nombre, datos_num[0],datos_num[2], datos_num[3], datos_num[4], datos_num[6], datos_num[5], datos_num[7], datos_num[1]]
            all_data.append(reorganizado)
        time.sleep(0.5)
        
       
    driver.quit()
    return all_data

def data_to_pdf(scraped_data):
    
    workbook = Workbook()
    sheet = workbook.active

    current_month_year = datetime.now().strftime("%B %Y")

    sheet.cell(row=482, column=1, value=current_month_year)
    sheet.cell(row=482, column=2, value="NEUC")
    
    for row_index, row_values in enumerate(scraped_data, start=482):
        for col_index, cell_value in enumerate(row_values, start=3):
            sheet.cell(row=row_index, column=col_index, value=cell_value)
    
    workbook.save('Neu.xlsx')

scraped_data = scrape_neu_com_co_tarifas()
data_to_pdf(scraped_data)
