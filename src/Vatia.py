import time
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from bs4 import BeautifulSoup
from openpyxl import Workbook
from selenium.webdriver.common.by import By
from datetime import datetime

driver = webdriver.Chrome()

def scrape_vatia_com_co_tarifas():
    driver.get("https://vatia.com.co/tarifas-costo-unitario-mercado-regulado/")
    time.sleep(5)
    all_data = {}
    html_source = driver.page_source
    soup = BeautifulSoup(html_source, 'html.parser')

    try:
        celdas = soup.find('table').find('tbody').find_all('td')
        for i in range(0, len(celdas), 17):
            row_data = celdas[i:i+17]
            data_list = [cell.text.strip() for cell in row_data]
            if data_list[0].startswith('202402'):
                key = data_list[1] 
                filtered_data = data_list[1:3] + data_list[4:11]
                if key in all_data:
                    all_data[key].append(filtered_data)
                else:
                    all_data[key] = [filtered_data]

        for key, data_list in all_data.items():
            for data_row in data_list:
                data_row[0], data_row[1], data_row[2], data_row[3], data_row[4], data_row[5], data_row[6], data_row[7], data_row[8] = data_row[0], data_row[1], data_row[2], data_row[5], data_row[6], data_row[4], data_row[3], data_row[7], data_row[8] 
    
    except Exception as e:
        print("Error durante la extracción de datos:", e)
    print(all_data)
    driver.quit()
    return all_data


def data_to_pdf(scraped_data):
    
    workbook = Workbook()
    sheet = workbook.active

    current_month_year = datetime.now().strftime("%B %Y")

    sheet.cell(row=1, column=1, value=current_month_year)
    
    row_index = 1 
    for company, data_list in scraped_data.items():
        first_row_for_company = True
        for data_row in data_list:
            if first_row_for_company:
                sheet.cell(row=row_index, column=3, value=company)
                first_row_for_company = False
            for col_index, cell_value in enumerate(data_row, start=2):
                sheet.cell(row=row_index, column=col_index, value=cell_value)
            row_index += 1
    
    workbook.save('Vatia.xlsx')

scraped_data = scrape_vatia_com_co_tarifas()
data_to_pdf(scraped_data)
